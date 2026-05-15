# -*- coding: utf-8 -*-
# 金蝶云星空经营数据（多表）导出脚本
#
# 说明：
# - 本文件是一个自包含版本，用于随 Skill 目录一起发布
# - 配置位于同目录 `config.py`

import requests
import json
import sys
import os
import copy
import re
import importlib.util
from datetime import datetime, timedelta
from dateutil.relativedelta import relativedelta

try:
    from config import KINGDEE_CONFIG, WECHAT_CONFIG
except ModuleNotFoundError:
    config_example_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "config.example.py")
    spec = importlib.util.spec_from_file_location("config_example", config_example_path)
    config_example = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(config_example)
    KINGDEE_CONFIG = config_example.KINGDEE_CONFIG
    WECHAT_CONFIG = config_example.WECHAT_CONFIG

try:
    import pandas as pd
except ModuleNotFoundError:
    print("缺少依赖库 pandas。请执行：python -m pip install -r requirements.txt")
    raise


class SalesDataExporter:
    """销售单据数据导出器"""

    def __init__(self, start_date=None, end_date=None, no_wechat=False, org_numbers=None, only=None, extra_fields=None):
        self.kingdee_config = KINGDEE_CONFIG
        self.wechat_webhook = (WECHAT_CONFIG or {}).get("webhook", "")
        self.session = requests.Session()
        self.session.headers.update({"Content-Type": "application/json"})
        self.base_url = self.kingdee_config["base_url"] + "/k3cloud/"

        self.no_wechat = no_wechat
        self.only = self._normalize_only(only)
        self.official_fields_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "官方字段说明")
        self.official_field_cache = {}
        self.requested_extra_fields = self._parse_extra_fields(extra_fields)
        self.requested_org_numbers = self._parse_org_numbers(org_numbers) if org_numbers else None

        # 日期范围：默认 1-6 号导出上月整月，7 号及以后导出当月 1 号到今天
        if start_date and end_date:
            self.start_date = start_date
            self.end_date = end_date
            start_dt = datetime.strptime(start_date, "%Y-%m-%d")
            self.period_name = start_dt.strftime("%Y年%m月")
            self.year = start_dt.year
            self.period = start_dt.month
        else:
            today = datetime.now()
            if today.day <= 6:
                first_day_this_month = today.replace(day=1)
                last_day_last_month = first_day_this_month - timedelta(days=1)
                first_day_last_month = last_day_last_month.replace(day=1)
                self.start_date = first_day_last_month.strftime("%Y-%m-%d")
                self.end_date = last_day_last_month.strftime("%Y-%m-%d")
                self.period_name = last_day_last_month.strftime("%Y年%m月")
                self.year = last_day_last_month.year
                self.period = last_day_last_month.month
            else:
                first_day_this_month = today.replace(day=1)
                self.start_date = first_day_this_month.strftime("%Y-%m-%d")
                self.end_date = today.strftime("%Y-%m-%d")
                self.period_name = today.strftime("%Y年%m月")
                self.year = today.year
                self.period = today.month

        # 默认不写死任何组织编码/公司名；登录后按 --org 或系统组织动态解析
        self.target_settle_org_numbers = []
        self.default_settle_org_id_map = {}
        self.sale_org_numbers = []
        self.inventory_org_number = None
        self.account_book_number = ""
        self.bill_configs = self._build_bill_configs()
        self.report_configs = self._build_report_configs()
        self._attach_official_fields_to_configs()

        print(f"查询{self.period_name}数据，日期范围: {self.start_date} 至 {self.end_date}")

    def _build_bill_configs(self):
        sale_org_filter = self._build_org_filter("FSaleOrgId.FNumber", self.sale_org_numbers)
        settle_org_filter = self._build_org_filter("FSETTLEORGID.FNumber", self.target_settle_org_numbers)
        pay_org_filter = self._build_org_filter("FPAYORGID.FNumber", self.target_settle_org_numbers)
        purchase_org_filter = self._build_org_filter("FPurchaseOrgId.FNumber", self.target_settle_org_numbers)

        return [
            {
                "form_id": "SAL_OUTSTOCK",
                "bill_name": "销售出库单",
                "field_keys": "FBillNo,FDate,FCustomerID.FName,FBillTypeID.FName,FSaleDeptID.FName,FSalesManID.FName,FNote,FMaterialID.FName,FStockID.FName,FEntryTaxAmount,FAmount,FAllAmount,FEntryCostAmount",
                "filter_string": f"{sale_org_filter} AND FDate>='{self.start_date}' AND FDate<='{self.end_date}' AND FDocumentStatus='C'",
                "columns": ["单据编号", "日期", "客户", "单据类型", "销售部门", "销售员", "备注", "物料名称", "仓库", "税额", "金额", "价税合计", "总成本"],
            },
            {
                "form_id": "SAL_SaleOrder",
                "bill_name": "销售订单",
                "field_keys": "FDate,FBillTypeID.FName,FBillNo,FDocumentStatus,FCustId.FName,FSaleDeptId.FName,FSalerId.FName,FCreatorId.FName,FCloseStatus,FMaterialId.FNumber,FMaterialName,FUnitID.FName,FQty,FPrice,FEntryTaxAmount,FAllAmount,FDeliveryDate",
                "filter_string": f"{sale_org_filter} AND FDate>='{self.start_date}' AND FDate<='{self.end_date}'",
                "columns": ["日期", "单据类型", "单据编号", "单据状态", "客户", "销售部门", "销售员", "创建人", "关闭状态", "物料编码", "物料名称", "销售单位", "销售数量", "单价", "税额", "价税合计", "要货日期"],
            },
            {
                "form_id": "SAL_RETURNSTOCK",
                "bill_name": "销售退货单",
                "field_keys": "FBillNo,FDate,FRetcustId.FName,FBillTypeID.FName,FSaledeptid.FName,FSalesManId.FName,FHeadNote,FMaterialId.FName,FStockId.FName,FEntryTaxAmount,FAmount,FAllAmount,FEntryCostAmount",
                "filter_string": f"{sale_org_filter} AND FDate>='{self.start_date}' AND FDate<='{self.end_date}' AND FDocumentStatus='C'",
                "columns": ["单据编号", "日期", "退货客户", "单据类型", "销售部门", "销售员", "备注", "物料名称", "仓库", "税额", "金额", "价税合计", "总成本"],
            },
            {
                "form_id": "AR_receivable",
                "bill_name": "应收单",
                "field_keys": "FBillTypeID.FName,FBillNo,FDATE,FCUSTOMERID.FName,FSALEORGID.FName,FSALEDEPTID.FName,FMATERIALID.FName,FTAXAMOUNTFOR_D,FNoTaxAmountFor_D,FALLAMOUNTFOR_D,FCreatorId.FName,FAR_Remark",
                "filter_string": f"{self._build_org_filter('FSALEORGID.FNumber', self.target_settle_org_numbers)} AND FDATE>='{self.start_date}' AND FDATE<='{self.end_date}' AND FDocumentStatus='C'",
                "columns": ["单据类型", "单据编号", "业务日期", "客户", "销售组织", "销售部门", "物料名称", "税额", "不含税金额", "价税合计", "创建人", "备注"],
            },
            {
                "form_id": "AP_Payable",
                "bill_name": "应付单",
                "field_keys": "FBillTypeID.FName,FDATE,FSUPPLIERID.FName,FBillNo,FSETTLEORGID.FName,FPURCHASEDEPTID.FName,FMATERIALID.FName,FCostName,FCOSTDEPARTMENTID.FName,FCreatorId.FName,FTAXAMOUNTFOR_D,FNOTAXAMOUNT_D,FALLAMOUNT_D,FAP_Remark",
                "filter_string": f"{settle_org_filter} AND FDATE>='{self.start_date}' AND FDATE<='{self.end_date}' AND FDOCUMENTSTATUS='C'",
                "columns": ["单据类型", "业务日期", "供应商", "单据编号", "结算组织", "采购部门", "物料名称", "费用项目名称", "费用承担部门", "创建人", "税额", "不含税金额本位币", "价税合计本位币", "备注"],
            },
            {
                "form_id": "PUR_PurchaseOrder",
                "bill_name": "采购订单",
                "field_keys": "FBillNo,FDate,FSupplierId.FName,FDocumentStatus,FPurchaseOrgId.FName,FPurchaserId.FName,FCreatorId.FName,FCloseStatus,F_TJKT_ChangeReason_re5,FMaterialId.FNumber,FMaterialName,FUnitId.FName,FQty,FDeliveryDate,FPrice,FEntryTaxAmount,FAllAmount,FGiveAway",
                "filter_string": f"{purchase_org_filter} AND FDate>='{self.start_date}' AND FDate<='{self.end_date}'",
                "columns": ["单据编号", "采购日期", "供应商", "单据状态", "采购组织", "采购员", "创建人", "关闭状态", "摘要", "物料编码", "物料名称", "采购单位", "采购数量", "交货日期", "单价", "税额", "价税合计", "是否赠品"],
            },
            {
                "form_id": "ER_ExpenseRequest",
                "bill_name": "费用申请单",
                "field_keys": "FBillNo,FDate,FStaffID.FName,FDeptID.FName,FOrgID.FName,FExpenseItemID.FName,FReason,FIsBorrow,FDocumentStatus,FCloseStatus,FOrgAmount,FCheckedOrgAmount",
                "filter_string": f"{self._build_org_filter('FOrgID.FNumber', self.target_settle_org_numbers)} AND FDate>='{self.start_date}' AND FDate<='{self.end_date}'",
                "columns": ["单据编号", "申请日期", "申请人", "申请部门", "申请组织", "费用项目", "事由", "申请借款", "单据状态", "关闭状态", "申请金额", "核定金额"],
            },
            {
                "form_id": "ER_ExpReimbursement",
                "bill_name": "费用报销单",
                "field_keys": "FBillTypeID.FName,FRealPay,FBillNo,FCausa,FDate,FProposerID.FName,FRequestDeptID.FName,FOrgID.FName,FRequestType,FExpID.FName,FDocumentStatus,FExpenseAmount,FRequestAmount,FPayedAmount,FRefundedAmount,FBorrowAmount,FOffsetAmount,FReimbNotPayAmount",
                "filter_string": f"{self._build_org_filter('FOrgID.FNumber', self.target_settle_org_numbers)} AND FDate>='{self.start_date}' AND FDate<='{self.end_date}' AND FCancelStatus='A'",
                "columns": ["单据类型", "实报实付", "单据编号", "事由", "申请日期", "申请人", "申请部门", "申请组织", "退款/付款", "费用项目", "单据状态", "申请报销金额", "申请退/付款金额", "已付款金额", "已退款金额", "冲借款金额", "冲销金额", "报销未付款金额"],
            },
            {
                "form_id": "ER_ExpenseRequest_Travel",
                "bill_name": "出差申请单",
                "field_keys": "FBillNo,FDate,FReason,FExpenseItemID.FName,FStaffID.FName,FDeptID.FName,FOrgID.FName,FIsBorrow,FOrgAmount,FDocumentStatus,FCheckedOrgAmount,FCloseStatus",
                "filter_string": f"{self._build_org_filter('FOrgID.FNumber', self.target_settle_org_numbers)} AND FDate>='{self.start_date}' AND FDate<='{self.end_date}' AND FCancelStatus='A'",
                "columns": ["单据编号", "申请日期", "事由", "费用项目", "申请人", "申请部门", "申请组织", "申请借款", "申请金额", "单据状态", "核定金额", "关闭状态"],
            },
            {
                "form_id": "ER_ExpReimbursement_Travel",
                "bill_name": "差旅费报销单",
                "field_keys": "FBillTypeID.FName,FRealPay,FBillNo,FCausa,FDate,FProposerID.FName,FRequestDeptID.FName,FOrgID.FName,FRequestType,FExpID.FName,FDocumentStatus,FExpenseAmount,FExpenseOrgId.FName,FRequestAmount,FPayedAmount,FReimbNotPayAmount,FRemark",
                "filter_string": f"{self._build_org_filter('FOrgID.FNumber', self.target_settle_org_numbers)} AND FDate>='{self.start_date}' AND FDate<='{self.end_date}' AND FCancelStatus='A'",
                "columns": ["单据类型", "实报实付", "单据编号", "事由", "申请日期", "申请人", "申请部门", "申请组织", "退款/付款", "费用项目", "单据状态", "申请报销金额", "费用承担组织", "申请退/付款金额", "已付款金额", "报销未付款金额", "备注"],
            },
            {
                "form_id": "CN_PAYAPPLY",
                "bill_name": "付款申请单",
                "field_keys": "FBILLTYPEID.FName,FBillNo,FDATE,FCONTACTUNIT.FName,FCURRENCYID.FName,FPAYAMOUNTFOR_H,FAPPLYAMOUNTFOR_H,FSETTLECUR.FName,FSETTLEORGID.FName,FCREATORID.FName,FDEPARTMENT.FName,FDOCUMENTSTATUS,FCLOSESTATUS,FPAYPURPOSEID.FName,FENDDATE,FCOSTID.FName,FDescription",
                "filter_string": f"{self._build_org_filter('FSETTLEORGID.FNumber', self.target_settle_org_numbers)} AND FDATE>='{self.start_date}' AND FDATE<='{self.end_date}' AND FCANCELSTATUS='A'",
                "columns": ["单据类型", "单据编号", "申请日期", "往来单位", "币别", "应付金额", "申请付款金额", "结算币别", "结算组织", "创建人", "部门", "单据状态", "关闭状态", "付款用途", "到期日", "费用项目", "备注"],
            },
            {
                "form_id": "AP_PAYBILL",
                "bill_name": "付款单",
                "field_keys": "FBillTypeID.FName,FBillNo,FDATE,FCONTACTUNITTYPE,FCONTACTUNIT.FName,FREMARK,FSETTLETYPEID.FName,FPURPOSEID.FName,FPAYORGID.FName,FCOSTID.FName,FEXPENSEDEPTID_E.FName,FHANDLINGCHARGEFOR,FREALPAYAMOUNTFOR_D",
                "filter_string": f"{pay_org_filter} AND FDATE>='{self.start_date}' AND FDATE<='{self.end_date}' AND FDOCUMENTSTATUS='C'",
                "columns": ["单据类型", "单据编号", "业务日期", "往来单位类型", "往来单位", "备注", "结算方式", "付款用途", "付款组织", "费用项目", "费用承担部门", "手续费", "表体-实付金额"],
            },
            {
                "form_id": "AR_RECEIVEBILL",
                "bill_name": "收款单",
                "field_keys": "FBillTypeID.FName,FBillNo,FDATE,FCONTACTUNITTYPE,FSETTLETYPEID.FName,FPURPOSEID.FName,FPAYORGID.FName,FSALEDEPTID.FName,FCONTACTUNIT.FName,FREMARK,FHANDLINGCHARGEFOR,FREALRECAMOUNTFOR_D",
                "filter_string": f"{pay_org_filter} AND FDATE>='{self.start_date}' AND FDATE<='{self.end_date}' AND FDOCUMENTSTATUS='C'",
                "columns": ["单据类型", "单据编号", "业务日期", "往来单位类型", "结算方式", "收款用途", "收款组织", "销售部门", "往来单位", "备注", "手续费", "表体-实收金额"],
            },
            {
                "form_id": "AP_REFUNDBILL",
                "bill_name": "付款退款单",
                "field_keys": "FBillTypeID.FName,FBillNo,FDATE,FCONTACTUNIT.FName,FPAYUNIT.FName,FSETTLETYPEID.FName,FPURPOSEID.FName,FREALREFUNDAMOUNTFOR_D,FPAYORGID.FName,FDepartment.FName,FEXPENSEDEPTID_E.FName,FCOSTID.FName,FREMARK",
                "filter_string": f"{pay_org_filter} AND FDATE>='{self.start_date}' AND FDATE<='{self.end_date}' AND FDOCUMENTSTATUS='C'",
                "columns": ["单据类型", "单据编号", "业务日期", "往来单位", "付款单位", "结算方式", "原付款用途", "表体-实退金额", "付款组织", "部门", "费用承担部门", "费用项目", "备注"],
            },
            {
                "form_id": "AR_REFUNDBILL",
                "bill_name": "收款退款单",
                "field_keys": "FBillTypeID.FName,FBillNo,FDATE,FCONTACTUNIT.FName,FSETTLETYPEID.FName,FPURPOSEID.FName,FREFUNDAMOUNTFOR_E,FPAYORGID.FName,FSALEDEPTID.FName,FREMARK",
                "filter_string": f"{pay_org_filter} AND FDATE>='{self.start_date}' AND FDATE<='{self.end_date}' AND FDOCUMENTSTATUS='C'",
                "columns": ["单据类型", "单据编号", "业务日期", "往来单位", "结算方式", "原收款用途", "表体-实退金额", "付款组织", "销售部门", "备注"],
            },
            {
                "form_id": "AP_OtherPayable",
                "bill_name": "其他应付单",
                "field_keys": "FBillTypeID.FName,FBillNo,FDATE,FCONTACTUNITTYPE,FCONTACTUNIT.FName,FTOTALAMOUNTFOR_H,FCOSTNAME,FDEPARTMENTID.FName,FSETTLEORGID.FName,FCreatorId.FName",
                "filter_string": f"{settle_org_filter} AND FDATE>='{self.start_date}' AND FDATE<='{self.end_date}' AND FDOCUMENTSTATUS='C'",
                "columns": ["单据类型", "单据编号", "业务日期", "往来单位类型", "往来单位", "总金额", "费用项目名称", "申请部门", "结算组织", "创建人"],
            },
            {
                "form_id": "AR_OtherRecAble",
                "bill_name": "其他应收单",
                "field_keys": "FBillTypeID.FName,FBillNo,FDATE,FCONTACTUNITTYPE,FCONTACTUNIT.FName,FAMOUNTFOR,FCOSTNAME,FDEPARTMENTID.FName,FSETTLEORGID.FName,FCreatorId.FName",
                "filter_string": f"{settle_org_filter} AND FDATE>='{self.start_date}' AND FDATE<='{self.end_date}' AND FDOCUMENTSTATUS='C'",
                "columns": ["单据类型", "单据编号", "业务日期", "往来单位类型", "往来单位", "总金额", "费用项目名称", "申请部门", "结算组织", "创建人"],
            },
            {
                "form_id": "AP_AdjustExchangeRate",
                "bill_name": "应付调汇单",
                "field_keys": "FBillNo,FCONTACTUNITTYPE,FCONTACTUNIT.FName,FBUSINESSDEPTID.FName,FDATE,FADJEXCAMOUNT",
                "filter_string": f"{settle_org_filter} AND FDATE>='{self.start_date}' AND FDATE<='{self.end_date}' AND FDOCUMENTSTATUS='C'",
                "columns": ["单据编号", "往来单位类型", "往来单位", "业务部门", "业务日期", "调汇金额"],
            },
        ]

    def _build_report_configs(self):
        inventory_org_number = self.inventory_org_number or (self.target_settle_org_numbers[0] if self.target_settle_org_numbers else "")

        return [
            {
                "form_id": "AP_SumReport",
                "report_name": "应付款汇总表",
                "field_keys": "FCONTACTUNITNUMBER,FCONTACTUNITNAME,FSETTLEORGNAME,FINITAMOUNT,FAMOUNT,FREALAMOUNT,FOFFAMOUNT,FLEFTAMOUNT",
                "model": {
                    "FCONTACTUNITTYPE": "供应商",
                    "FUSEDATE": "true",
                    "FBeginDate": self.start_date,
                    "FEndDate": self.end_date,
                    "FSettleOrgLst": "1",
                    "FOutSettle": "true",
                    "FInSettle": "false",
                },
                "columns": ["往来单位编码", "往来单位名称", "结算组织", "(本位币)期初余额", "(本位币)本期应付", "(本位币)本期付款", "(本位币)本期冲销额", "(本位币)期末余额"],
            },
            {
                "form_id": "AR_SumReport",
                "report_name": "应收款汇总表",
                "field_keys": "FCONTACTUNITNUMBER,FCONTACTUNITNAME,FSETTLEORGNAME,FINITAMOUNTFOR,FAMOUNTFOR,FREALAMOUNTFOR,FOFFAMOUNTFOR,FLEFTAMOUNTFOR",
                "model": {
                    "FCONTACTUNITTYPE": "客户",
                    "FUSEDATE": "true",
                    "FBeginDate": self.start_date,
                    "FEndDate": self.end_date,
                    "FSettleOrgLst": "",
                    "FOutSettle": "true",
                    "FInSettle": "false",
                },
                "columns": ["往来单位编码", "往来单位名称", "结算组织", "(原币)期初余额", "(原币)本期应收", "(原币)本期收款", "(原币)本期冲销额", "(原币)期末余额"],
            },
            {
                "form_id": "HS_INOUTSTOCKSUMMARYRPT",
                "report_name": "存货收发存汇总表",
                "field_keys": "FMATERIALBASEID,FMATERIALNAME,FMATERIALGROUP,FSTOCKId,FINITQty,FINITPrice,FINITAMOUNT,FRECEIVEQty,FRECEIVEPrice,FRECEIVEAmount,FSENDQty,FSENDPrice,FSENDAmount,FENDQty,FENDPrice,FENDAmount",
                "model": {
                    "FACCTGSYSTEMID": {"FNumber": "KJHSTX01_SYS"},
                    "FACCTGORGID": {"FNumber": inventory_org_number},
                    "FACCTPOLICYID": {"FNumber": "KJZC01_SYS"},
                    "FYear": str(self.year),
                    "FPeriod": str(self.period),
                    "FENDYEAR": str(self.year),
                    "FEndPeriod": str(self.period),
                    "FCOMBOTotalType": "不汇总",
                    "FDimType": "FMATERIALID,FSTOCKID",
                    "FIsDisplayPeriod": True,
                },
                "columns": ["物料编码", "物料名称", "物料分组", "仓库", "期初数量", "期初单价", "期初金额", "收入数量", "收入单价", "收入金额", "发出数量", "发出单价", "发出金额", "期末数量", "期末单价", "期末金额"],
            },
            {
                "form_id": "HS_NoDimInOutStockDetailRpt",
                "report_name": "存货收发存明细表",
                "field_keys": "FPERIOD,FBILLDATE,FBILLNO,FBUSINESSTYPE,FBillFormName,FMATERIALID,FMATERIALNAME,FRECEIVEQty,FRECEIVEPrice,FRECEIVEAmount,FSENDQty,FSENDPrice,FSENDAmount,FENDQty,FENDPrice,FENDAmount",
                "model": {
                    "FACCTGSYSTEMID": {"FNumber": "KJHSTX01_SYS"},
                    "FACCTGORGID": {"FNumber": inventory_org_number},
                    "FACCTPOLICYID": {"FNumber": "KJZC01_SYS"},
                    "FYear": str(self.year),
                    "FENDYEAR": str(self.year),
                    "FPeriod": str(self.period),
                    "FEndPeriod": str(self.period),
                },
                "columns": ["期间", "单据日期", "单据编号", "业务类型", "单据类型", "物料编码", "物料名称", "收入数量", "收入单价", "收入金额", "发出数量", "发出单价", "发出金额", "期末数量", "期末单价", "期末金额"],
            },
            {
                "form_id": "CN_FundPositionReport",
                "report_name": "资金头寸表",
                "field_keys": "FRowTypeName,FBankName,FBankAcctName,FBankAcctNo,FPAYORGNAME,FINNERACCTNAME,FINNERACCTNO,FForCurrencyName,FForLastBal,FForTodayIn,FForTodayOut,FForTodayBal,FLocalCurrencyName,FLocalLastBal,FLocalTodayIn,FLocalTodayOut,FLocalTodayBal,FInCount,FOutCount",
                "model": {
                    "FOrgId": [{"FNumber": number} for number in self.target_settle_org_numbers],
                    "FStartDate": f"{self.start_date} 00:00:00",
                    "FEndDate": f"{self.end_date} 00:00:00",
                    "FNotAudit": False,
                    "FInNOut": True,
                    "FMyCurrency": False,
                    "FCurrencySubTotal": False,
                    "FSettleOrgBox": False,
                    "FPAYORGIDBOX": True,
                    "FMyCurrencySum": False,
                    "FMyPayOrg": False,
                    "FGroupCash": False,
                    "FOrgOrAccount": "0",
                    "FAllCashAccount": True,
                    "FAllBankAccount": False,
                    "FIsShowCancelBankAcnt": False,
                    "FINCLUDEEMPTY": False,
                },
                "columns": ["资金类别", "银行", "账户名称", "银行账号", "收付组织", "内部账户名称", "内部账户", "原币币别", "原币期初余额", "原币本日收入", "原币本日支出", "原币本日余额", "本位币币别", "本位币期初余额", "本位币本日收入", "本位币本日支出", "本位币本日余额", "收入笔数", "支出笔数"],
            },
            {
                "form_id": "SAL_OutStockInvoiceRpt",
                "report_name": "销售出库开票跟踪表",
                "field_keys": "FSALEORGNAME,FBILLNO,FBILLTYPENAME,FDate,FSALESNAME,FCUSTOMERNAME,FMATERIALNAME,FREALQTY,FPrice,FALLAMOUNT,FISFREE,FRECQTY,FRECAMOUNT,FWriteOffAmount,FINVOECEQTY,FINVOECEAMOUNT,FRECEIPTAMOUNT,FJSWRITEOFFAMOUNT,FChargeOffAmount",
                "org_id_model_field": "FSaleOrgId",
                "model": {
                    "FSaleOrgId": "",
                    "FMoneyType": {"FNumber": ""},
                    "FStartDate": self.start_date,
                    "FEndDate": self.end_date,
                    "FCustomerFrom": {"FNumber": ""},
                    "FCustomerTo": {"FNumber": ""},
                    "FSaleDeptFrom": {"FNUMBER": ""},
                    "FSaleDeptTo": {"FNUMBER": ""},
                    "FMaterialFrom": {"FNumber": ""},
                    "FMaterialTo": {"FNumber": ""},
                    "FFormStatus": "C",
                    "FIsIncludeSerMat": "false",
                    "FSuite": "",
                    "FSettleOrgList": "",
                },
                "columns": ["销售组织", "单据编号", "单据类型", "日期", "销售员", "客户名称", "物料名称", "数量", "单价", "金额", "是否赠品", "应收数量", "应收金额", "调整金额", "开票数量", "开票金额", "结算金额", "结算调整金额", "特殊冲销金额"],
            },
            {
                "form_id": "PUR_PurchaseOrderDetailRpt",
                "report_name": "采购订单执行明细表",
                "field_keys": "FPurchaseOrgId,FBillNo,FDate,FSUPPLIERNAME,FMATERIALNAME,FDELIVERYDATE,FCurrencyId,FOrderQty,FOrderAmount,FReceiveQty,FReceiveAmount,FImportQty,FImportAmount,FReturnQty,FReturnAmount,FPAYQTY,FPAYAMOUNT,FPREINVOICEQTY,FPREINVOICEAMOUNT,FINVOICEQTY,FINVOICEAMOUNT,FRECPAYBILLAMOUNT,FPAYBILLAMOUNT,FSETADJAMOUNT,FPAYWRITOFFAMOUNT,FSPEWOFFAMOUNT",
                "org_id_model_field": "FPurchaseOrgIdList",
                "model": {
                    "FPurchaseOrgIdList": "",
                    "FOrderStartDate": self.start_date,
                    "FOrderEndDate": self.end_date,
                    "FBeginSupplierId": {"FNumber": ""},
                    "FEndSupplierId": {"FNumber": ""},
                    "FBeginBillNumber": "",
                    "FEndBillNumber": "",
                    "FBeginMaterialId": {"FNumber": ""},
                    "FEndMaterialId": {"FNumber": ""},
                    "FBeginPurchaser": {"FNumber": ""},
                    "FEndFPurchaser": {"FNumber": ""},
                    "FBusinessType": "",
                    "FDocumentStatus": "C",
                    "FLineStatus": "A",
                },
                "columns": ["采购组织", "订单编号", "日期", "供应商名称", "物料名称", "交货日期", "结算币别", "订货数量", "价税合计", "收料数量", "收料金额", "入库数量", "入库金额", "退料数量", "退料金额", "应付数量", "应付金额", "先开票数量", "先开票金额", "开票数量", "开票金额", "预付金额", "已结算金额", "结算调整金额", "付款核销金额", "特殊冲销金额"],
            },
            self._build_kds_report_config("财务报表", "BBMB0001", inventory_org_number),
            {
                "form_id": "GL_RPT_AccountBalance",
                "report_name": "科目余额表",
                "field_keys": "FBALANCEID,FBALANCENAME,FDETAILNUMBER,FDETAILNAME,FBEGINDEBITLOCAL,FBEGINCREDITLOCAL,FDEBITLOCAL,FCREDITLOCAL,FYTDDEBITLOCAL,FYTDCREDITLOCAL,FENDDEBITLOCAL,FENDCREDITLOCAL",
                "scheme_id": "69c396dfde2072",
                "model": {
                    "FACCTBOOKID": {"FNumber": self.account_book_number},
                    "FCURRENCY": "0",
                    "FSTARTYEAR": str(self.year),
                    "FSTARTPERIOD": str(self.period),
                    "FENDYEAR": str(self.year),
                    "FENDPERIOD": str(self.period),
                    "FBALANCELEVEL": "3",
                    "FSHOWDETAIL": True,
                    "FFORBIDBALANCE": True,
                    "FNOTPOSTVOUCHER": True,
                    "FDEBITORCREDIT": False,
                    "FBALANCEZERO": True,
                    "FNOBUSINESS": False,
                    "FPERIODNOBALANCE": True,
                    "FYEARNOBALANCE": True,
                    "FSHOWFULLNAME": True,
                    "FDETAILSHOWACCT": True,
                    "FSHOWDETAILONLY": False,
                    "FEXCLUDEADJUSTVCH": False,
                    "FFLEXDEBITORCREDIT": False,
                    "FSHOWFLEXBYCOL": False,
                },
                "columns": ["科目编码", "科目名称", "核算维度编码", "核算维度名称", "期初余额-本位币（借）", "期初余额-本位币（贷）", "本期发生-本位币（借）", "本期发生-本位币（贷）", "本年累计-本位币（借）", "本年累计-本位币（贷）", "期末余额-本位币（借）", "期末余额-本位币（贷）"],
            },
        ]

    def _build_kds_report_config(self, report_name, report_number, org_number):
        financial_report_config = (self.kingdee_config or {}).get("financial_report", {}) or {}
        cycle_type = int(financial_report_config.get("CycleType", 4))
        return {
            "form_id": "KDS_ReportData",
            "report_name": report_name,
            "api_type": "kds_report",
            "model": {
                "ReportType": int(financial_report_config.get("ReportType", 1)),
                "ReportNumber": financial_report_config.get("ReportNumber", report_number),
                "AcctSystemNumber": financial_report_config.get("AcctSystemNumber", "KJHSTX01_SYS"),
                "AcctPolicyNumber": financial_report_config.get("AcctPolicyNumber", "KJZC01_SYS"),
                "OrgNumber": org_number,
                "CurrencyNumber": financial_report_config.get("CurrencyNumber", "PRE001"),
                "CurrUnitNumber": financial_report_config.get("CurrUnitNumber", "JEDW01_SYS"),
                "CycleType": cycle_type,
                "Year": int(financial_report_config.get("Year", self.year)),
                "Period": int(financial_report_config.get("Period", self._report_period_for_cycle(cycle_type))),
                "DataType": "Json",
                "ResultType": "0",
            },
            "columns": None,
        }

    def _report_period_for_cycle(self, cycle_type):
        if cycle_type == 5:
            return (self.period - 1) // 3 + 1
        if cycle_type == 6:
            return 1 if self.period <= 6 else 2
        if cycle_type == 7:
            return 1
        return self.period

    def _resolve_org_scope_after_login(self):
        if self.requested_org_numbers and self.requested_org_numbers != ["all"]:
            resolved_org_numbers = self.requested_org_numbers
        else:
            all_orgs = self.get_all_organizations()
            resolved_org_numbers = [o.get("number") for o in all_orgs if o.get("number")]
            resolved_org_numbers = [n for n in resolved_org_numbers if n]
            if self.requested_org_numbers == ["all"] and not resolved_org_numbers:
                raise RuntimeError("未能从金蝶查询到组织列表，无法使用 --org all")

        self.target_settle_org_numbers = resolved_org_numbers
        self.sale_org_numbers = list(resolved_org_numbers)
        self.inventory_org_number = "101" if "101" in resolved_org_numbers else (resolved_org_numbers[0] if resolved_org_numbers else None)
        self.account_book_number = self.resolve_account_book_number(self.inventory_org_number)
        self.bill_configs = self._build_bill_configs()
        self.report_configs = self._build_report_configs()
        self._attach_official_fields_to_configs()

    def _normalize_only(self, only):
        if not only:
            return None
        if isinstance(only, str):
            parts = [p.strip() for p in only.split(",") if p.strip()]
        else:
            parts = [str(p).strip() for p in (only or []) if str(p).strip()]
        lowered = {p.lower() for p in parts}
        return lowered or None

    def _parse_extra_fields(self, extra_fields):
        """解析 --fields，格式：导出项:字段1,字段2;另一个导出项:字段3。"""
        parsed = {}
        if not extra_fields:
            return parsed
        raw = str(extra_fields).strip()
        if not raw:
            return parsed
        for group in re.split(r"[;；]", raw):
            if not group.strip():
                continue
            if ":" in group:
                target, fields_part = group.split(":", 1)
            elif "：" in group:
                target, fields_part = group.split("：", 1)
            else:
                target, fields_part = "*", group
            target = target.strip().lower() or "*"
            fields = [f.strip() for f in re.split(r"[,，]", fields_part) if f.strip()]
            if fields:
                parsed.setdefault(target, []).extend(fields)
        return parsed

    def _official_doc_name_for_config(self, config):
        return config.get("official_doc") or config.get("bill_name") or config.get("report_name")

    def _load_official_fields(self, doc_name):
        if not doc_name:
            return {}
        if doc_name in self.official_field_cache:
            return self.official_field_cache[doc_name]

        path = os.path.join(self.official_fields_dir, f"{doc_name}.txt")
        fields = {}
        if os.path.exists(path):
            try:
                with open(path, "r", encoding="utf-8") as f:
                    for line in f:
                        match = re.search(r"\s*([^：:]+)[：:]\s*([A-Za-z_][A-Za-z0-9_.]*)", line)
                        if not match:
                            continue
                        name = match.group(1).strip()
                        key = match.group(2).strip()
                        if name and key:
                            fields[name] = key
                            fields[key.upper()] = key
            except Exception as e:
                print(f"  [WARN] 读取官方字段说明失败 {path}: {e}")

        self.official_field_cache[doc_name] = fields
        return fields

    def _attach_official_fields_to_configs(self):
        for config in self.bill_configs + self.report_configs:
            doc_name = self._official_doc_name_for_config(config)
            config["official_fields"] = self._load_official_fields(doc_name)

    def _append_extra_field_keys(self, config, field_keys):
        if not self.requested_extra_fields:
            return field_keys

        names = []
        for target in ("*", config.get("form_id", "").lower(), config.get("bill_name", "").lower(), config.get("report_name", "").lower()):
            names.extend(self.requested_extra_fields.get(target, []))
        if not names:
            return field_keys

        official_fields = config.get("official_fields") or {}
        keys = [k.strip() for k in str(field_keys or "").split(",") if k.strip()]
        key_set = {k.upper() for k in keys}
        for name in names:
            key = official_fields.get(name) or official_fields.get(name.upper())
            if not key and re.match(r"^[A-Za-z_][A-Za-z0-9_.]*$", name):
                key = name
            if not key:
                print(f"  [WARN] {config.get('bill_name') or config.get('report_name')} 未找到官方字段：{name}")
                continue
            if key.upper() not in key_set:
                keys.append(key)
                key_set.add(key.upper())
                print(f"  -> 已追加查询字段 {name}: {key}（默认不输出到Excel）")
        return ",".join(keys)

    def _parse_org_numbers(self, org_numbers):
        if isinstance(org_numbers, str):
            raw = org_numbers.strip()
            if raw.lower() == "all":
                return ["all"]
            parts = [p.strip() for p in raw.split(",") if p.strip()]
            return parts
        return [str(x).strip() for x in (org_numbers or []) if str(x).strip()]

    def _build_org_filter(self, field_name, org_numbers):
        org_numbers = [str(x).strip() for x in (org_numbers or []) if str(x).strip()]
        if not org_numbers:
            return "1=1"
        if len(org_numbers) == 1:
            return f"{field_name}='{org_numbers[0]}'"
        in_values = ",".join([f"'{n}'" for n in org_numbers])
        return f"{field_name} IN ({in_values})"

    def resolve_settle_org_ids_by_numbers(self, org_numbers, default_map=None):
        """根据组织编码解析结算组织内码，并返回逗号拼接字符串。"""
        default_map = default_map or {}
        org_id_map = dict(default_map)

        need_query_numbers = [number for number in org_numbers if number not in org_id_map]

        if need_query_numbers:
            url = self.base_url + "Kingdee.BOS.WebApi.ServicesStub.DynamicFormService.ExecuteBillQuery.common.kdsvc"
            in_values = ",".join([f"'{number}'" for number in need_query_numbers])
            data = {
                "FormId": "AP_Payable",
                "FieldKeys": "FSETTLEORGID,FSETTLEORGID.FNumber,FSETTLEORGID.FName",
                "FilterString": f"FSETTLEORGID.FNumber IN ({in_values})",
                "OrderString": "",
                "TopRowCount": 0,
                "StartRow": 0,
                "Limit": 2000,
            }
            payload = {"formid": "AP_Payable", "data": json.dumps(data, ensure_ascii=False)}

            try:
                response = self.session.post(url, json=payload, timeout=60)
                if response.status_code == 200:
                    result = response.json()
                    if isinstance(result, list):
                        for row in result:
                            if not isinstance(row, list) or len(row) < 2:
                                continue
                            org_id = str(row[0]).strip()
                            org_number = str(row[1]).strip()
                            if org_id and org_number and org_number not in org_id_map:
                                org_id_map[org_number] = org_id
                else:
                    print(f"  [WARN] 组织内码查询失败，状态码: {response.status_code}")
            except Exception as e:
                print(f"  [WARN] 组织内码查询异常: {e}")

        settle_org_ids = [org_id_map[number] for number in org_numbers if number in org_id_map and org_id_map[number]]
        settle_org_lst = ",".join(settle_org_ids)
        print(f"  -> 结算组织内码映射: {org_id_map}")
        print(f"  -> FSettleOrgLst: {settle_org_lst or '[空]'}")
        return settle_org_lst

    def resolve_account_book_number(self, org_number):
        """根据组织编码推断主账簿编码。"""
        if not org_number:
            return ""
        configured_books = (self.kingdee_config or {}).get("account_book_numbers", {}) or {}
        if org_number in configured_books:
            return str(configured_books[org_number]).strip()
        configured_book = (self.kingdee_config or {}).get("account_book_number", "")
        if configured_book:
            return str(configured_book).strip()

        url = self.base_url + "Kingdee.BOS.WebApi.ServicesStub.DynamicFormService.ExecuteBillQuery.common.kdsvc"
        data = {
            "FormId": "BD_AccountBook",
            "FieldKeys": "FBOOKID,FNumber,FName",
            "FilterString": "",
            "OrderString": "",
            "TopRowCount": 0,
            "StartRow": 0,
            "Limit": 200,
        }
        try:
            response = self.session.post(url, json={"formid": "BD_AccountBook", "data": json.dumps(data, ensure_ascii=False)}, timeout=60)
            rows = response.json() if response.status_code == 200 else []
            for row in rows:
                if not isinstance(row, list) or len(row) < 3:
                    continue
                number = str(row[1]).strip()
                name = str(row[2]).strip()
                if org_number in name:
                    print(f"  -> 科目余额表账簿: {number} {name}")
                    return number
            if len(rows) == 1 and isinstance(rows[0], list) and len(rows[0]) >= 2:
                return str(rows[0][1]).strip()
            non_group_books = [row for row in rows if isinstance(row, list) and len(row) >= 3 and "集团" not in str(row[2])]
            if len(non_group_books) == 1:
                return str(non_group_books[0][1]).strip()
        except Exception as e:
            print(f"  [WARN] 账簿编码查询异常: {e}")
        return ""

    def build_ap_sum_report_model(self, settle_org_lst):
        """构建应付款汇总表Model。"""
        return {
            "FAffiliation": {"FNAME": ""},
            "FSTARTYEAR": str(self.year),
            "FENDYEAR": str(self.year),
            "FCONTACTUNITTYPE": "BD_Supplier",
            "FSTARTPERIOD": str(self.period),
            "FENDPERIOD": str(self.period),
            "FCONTACTUNITFrom": {"FNumber": ""},
            "FUSEDATE": "true",
            "FSettleOrgLst": settle_org_lst,
            "FCONTACTUNITTo": {"FNumber": ""},
            "FAccountSystem": {"FNumber": ""},
            "FInSettle": "true",
            "FOutSettle": "true",
            "FUSEPERIOD": "",
            "FNoShowForNoLeft": "false",
            "FNoShowForNoOccur": "false",
            "FNOAUDIT": "false",
            "FEndDate": self.end_date,
            "FBeginDate": self.start_date,
            "FNoShowForBoth": "false",
            "FCurrencyFrom": [{"FNumber": ""}],
            "FIncludePayEvaluate": "false",
            "FIncludePayEvaluate_New": "false",
            "FDEFAULTACCTCALENDARID": 0,
            "FOnlyShowPayEvaluate": "false",
            "FOnlyShowPayEvaluate_New": "false",
            "FShowLocal": "true",
            "FShowSumLocal": "false",
            "FNoPrePayment": "false",
            "FOnlyShowPrePayment": "false",
            "FShowAmountInCost": "false",
            "FCONTACTUNITMUL": "",
            "FMULCONTACT": "false",
            "FPRESETBASE1": [{"FNumber": ""}],
            "FPRESETBASE2": [{"FNumber": ""}],
            "FGROUPSUPPLIER": "false",
            "FPERIODAMOUNT": -999999999,
            "FTOPERIODAMOUNT": 999999999,
            "FCheckPeriod": "false",
            "FShowMatchBill": "true",
            "FDateRadioGrp": "",
        }

    def build_ar_sum_report_model(self, settle_org_lst):
        """构建应收款汇总表Model。"""
        return {
            "FAffiliation": {"FNAME": ""},
            "FSTARTYEAR": str(self.year),
            "FENDYEAR": str(self.year),
            "FCONTACTUNITTYPE": "BD_Customer",
            "FSTARTPERIOD": str(self.period),
            "FENDPERIOD": str(self.period),
            "FCONTACTUNITFrom": {"FNumber": ""},
            "FUSEDATE": "true",
            "FSettleOrgLst": settle_org_lst,
            "FCONTACTUNITTo": {"FNumber": ""},
            "FAccountSystem": {"FNumber": ""},
            "FInSettle": "true",
            "FOutSettle": "true",
            "FUSEPERIOD": "",
            "FNoShowForNoLeft": "false",
            "FNoShowForNoOccur": "false",
            "FNOAUDIT": "false",
            "FEndDate": self.end_date,
            "FBeginDate": self.start_date,
            "FNoShowForBoth": "false",
            "FCurrencyFrom": [{"FNumber": ""}],
            "FIncludePayEvaluate": "false",
            "FIncludePayEvaluate_New": "false",
            "FDEFAULTACCTCALENDARID": 0,
            "FOnlyShowPayEvaluate": "false",
            "FOnlyShowPayEvaluate_New": "false",
            "FShowLocal": "true",
            "FGroupCustomer": "false",
            "FShowSumLocal": "false",
            "FNoPreReceive": "false",
            "FOnlyShowPreReceive": "false",
            "FCONTACTUNITMUL": "",
            "FMULCONTACT": "false",
            "FPRESETBASE1": [{"FNumber": ""}],
            "FPRESETBASE2": [{"FNumber": ""}],
            "FEXCLUDEB2CAR": "false",
            "FPERIODAMOUNT": -999999999,
            "FTOPERIODAMOUNT": 999999999,
            "FCheckPeriod": "false",
            "FShowMatchBill": "true",
            "FDateRadioGrp": "",
        }

    def login_kingdee(self):
        """登录金蝶云"""
        login_data = {
            "acctid": self.kingdee_config["acctid"],
            "username": self.kingdee_config["username"],
            "password": self.kingdee_config["password"],
            "lcid": 2052,
        }

        url = self.base_url + "Kingdee.BOS.WebApi.ServicesStub.AuthService.ValidateUser.common.kdsvc"

        try:
            response = self.session.post(url, json=login_data, timeout=30)
            if response.status_code == 200:
                result = response.json()
                if result.get("LoginResultType") == 1:
                    print("登录成功")
                    return True
            print("登录失败")
            return False
        except Exception as e:
            print(f"登录异常: {e}")
            return False

    def get_all_organizations(self):
        """
        尝试从金蝶查询所有组织（用于 --org all / --list-orgs）。
        说明：不同环境的组织基础资料 FormId / 字段名可能有差异，这里做多种兜底尝试。
        """
        if not self.login_kingdee():
            raise RuntimeError("登录失败，无法查询组织列表")

        candidates = [
            ("ORG_Organizations", "FOrgId,FNumber,FName,FForbidStatus"),
            ("ORG_Organizations", "FOrgID,FNumber,FName,FForbidStatus"),
            ("BD_Organization", "FOrgId,FNumber,FName,FForbidStatus"),
            ("ORG_Organizations", "FNumber,FName"),
            ("BD_Organization", "FNumber,FName"),
        ]

        url = self.base_url + "Kingdee.BOS.WebApi.ServicesStub.DynamicFormService.ExecuteBillQuery.common.kdsvc"

        for form_id, field_keys in candidates:
            data = {
                "formid": form_id,
                "data": json.dumps(
                    {
                        "FormId": form_id,
                        "FieldKeys": field_keys,
                        "FilterString": "",
                        "OrderString": "FNumber",
                        "TopRowCount": 0,
                        "StartRow": 0,
                        "Limit": 2000,
                    },
                    ensure_ascii=False,
                ),
            }
            try:
                resp = self.session.post(url, json=data, timeout=60)
                if resp.status_code != 200:
                    continue
                result = resp.json()
                if not isinstance(result, list):
                    continue

                rows = []
                for row in result:
                    if not isinstance(row, list) or len(row) < 2:
                        continue
                    number = str(row[1]).strip() if len(row) >= 2 else ""
                    name = str(row[2]).strip() if len(row) >= 3 else ""
                    org_id = str(row[0]).strip() if len(row) >= 1 else ""
                    rows.append({"id": org_id, "number": number, "name": name, "form_id": form_id})
                if any(r.get("number") for r in rows):
                    dedup = {}
                    for r in rows:
                        key = r.get("number") or r.get("name") or r.get("id")
                        if not key:
                            continue
                        dedup[key] = r
                    rows = list(dedup.values())
                    rows.sort(key=lambda x: x.get("number") or "")
                    return rows
            except Exception:
                continue

        raise RuntimeError("查询组织列表失败：请检查金蝶环境中组织基础资料的 FormId/字段是否不同")

    def get_bill_data_with_filter(self, form_id, field_keys, filter_string):
        """使用字段和过滤条件获取单据数据"""
        all_data = []
        start_row = 0
        limit = 2000

        url = self.base_url + "Kingdee.BOS.WebApi.ServicesStub.DynamicFormService.ExecuteBillQuery.common.kdsvc"

        print("  正在获取数据...")

        while True:
            data = {
                "formid": form_id,
                "data": json.dumps(
                    {
                        "FormId": form_id,
                        "FieldKeys": field_keys,
                        "FilterString": filter_string,
                        "OrderString": "",
                        "TopRowCount": 0,
                        "StartRow": start_row,
                        "Limit": limit,
                    }
                ),
            }

            try:
                response = self.session.post(url, json=data, timeout=60)

                if response.status_code == 200:
                    result = response.json()

                    if isinstance(result, list) and len(result) > 0:
                        all_data.extend(result)
                        print(f"  已获取 {len(all_data)} 条数据...")

                        if len(result) < limit:
                            break

                        start_row += limit
                    else:
                        break
                else:
                    print(f"  API请求失败: {response.status_code}")
                    print(f"  响应内容: {response.text[:500]}")
                    break

            except Exception as e:
                print(f"  获取数据异常: {e}")
                import traceback

                traceback.print_exc()
                break

        print(f"  共获取到 {len(all_data)} 条数据")
        return all_data

    def get_report_data(self, form_id, field_keys, model, scheme_id=""):
        """使用GetSysReportData接口获取报表数据"""
        all_data = []
        start_row = 0
        limit = 10000

        url = self.base_url + "Kingdee.BOS.WebApi.ServicesStub.DynamicFormService.GetSysReportData.common.kdsvc"

        print("  正在获取报表数据...")

        while True:
            data = {
                "FieldKeys": field_keys,
                "SchemeId": scheme_id,
                "StartRow": start_row,
                "Limit": limit,
                "IsVerifyBaseDataField": "true",
                "FilterString": [],
                "Model": model,
            }

            payload = {"formid": form_id, "data": json.dumps(data, ensure_ascii=False)}

            try:
                response = self.session.post(url, json=payload, timeout=120)

                if response.status_code == 200:
                    result = response.json()

                    if not result.get("Result", {}).get("IsSuccess", False):
                        error_msg = result.get("Result", {}).get("Message", "未知错误")
                        print(f"  API返回错误: {error_msg}")
                        print(f"  调试信息 - 完整响应: {json.dumps(result, ensure_ascii=False, indent=2)}")
                        break

                    rows = result.get("Result", {}).get("Rows", [])
                    row_count = result.get("Result", {}).get("RowCount", 0)

                    if rows and len(rows) > 0:
                        all_data.extend(rows)
                        print(f"  已获取 {len(all_data)} 条数据...")

                        if len(all_data) >= row_count or len(rows) < limit:
                            break

                        start_row += len(rows)
                    else:
                        break
                else:
                    print(f"  API请求失败: {response.status_code}")
                    print(f"  响应内容: {response.text[:500]}")
                    break

            except Exception as e:
                print(f"  获取报表数据异常: {e}")
                import traceback

                traceback.print_exc()
                break

        print(f"  共获取到 {len(all_data)} 条报表数据")
        return all_data

    def get_kds_report_data(self, model):
        """使用财务报表 GetReportData 接口获取报表数据。"""
        url = self.base_url + "Kingdee.BOS.KDS.ServiceFacade.ServicesStub.KDSReportAPIStub.GetReportData.common.kdsvc"
        payload = {"parameters": [json.dumps(model, ensure_ascii=False)]}

        print("  正在获取财务报表数据...")
        try:
            response = self.session.post(url, json=payload, timeout=120)
            if response.status_code != 200:
                print(f"  API请求失败: {response.status_code}")
                print(f"  响应内容: {response.text[:500]}")
                return []

            result = response.json()
            if isinstance(result, str):
                result = json.loads(result)
            if isinstance(result, dict) and str(result.get("status", "")).lower() in ("1", "false", "error"):
                print(f"  API返回错误: {result.get('message') or result}")
                return []
            return self._normalize_kds_report_result(result)
        except Exception as e:
            print(f"  获取财务报表数据异常: {e}")
            import traceback

            traceback.print_exc()
            return []

    def _normalize_kds_report_result(self, result):
        if isinstance(result, dict) and "result" in result:
            result = result.get("result")
            if isinstance(result, str):
                try:
                    result = json.loads(result)
                except Exception:
                    return [["报表结果", result]]

        spread_rows = self._extract_kds_spread_rows(result)
        if spread_rows:
            return spread_rows

        if isinstance(result, dict):
            for key in ("Rows", "rows", "Data", "data", "Result", "result"):
                value = result.get(key)
                if isinstance(value, list):
                    return value
            return [[k, v] for k, v in result.items()]
        if isinstance(result, list):
            return result
        return [["报表结果", result]]

    def _extract_kds_spread_rows(self, result):
        if not isinstance(result, dict):
            return []

        def walk(node):
            if isinstance(node, dict):
                if node.get("xtype") == "kdspread" and isinstance(node.get("data"), dict):
                    data = node["data"].get("data")
                    if isinstance(data, list) and data:
                        sheets = {}
                        for cells in data:
                            rows = self._kds_cells_to_rows(cells)
                            if not rows:
                                continue
                            title = str(rows[0][0]).strip() or f"报表{len(sheets) + 1}"
                            sheets[title[:31]] = rows
                        return sheets or []
                for value in node.values():
                    found = walk(value)
                    if found:
                        return found
            elif isinstance(node, list):
                for item in node:
                    found = walk(item)
                    if found:
                        return found
            return []

        cells = walk(result)
        if not cells:
            return []
        if isinstance(cells, dict):
            return cells

        return self._kds_cells_to_rows(cells)

    def _kds_cells_to_rows(self, cells):
        if not cells:
            return []
        max_row = max(int(cell[0]) for cell in cells if isinstance(cell, list) and len(cell) >= 3)
        max_col = max(int(cell[1]) for cell in cells if isinstance(cell, list) and len(cell) >= 3)
        rows = [["" for _ in range(max_col + 1)] for _ in range(max_row + 1)]
        for cell in cells:
            if not isinstance(cell, list) or len(cell) < 3:
                continue
            row_idx = int(cell[0])
            col_idx = int(cell[1])
            rows[row_idx][col_idx] = cell[2]

        while rows and all(str(v).strip() == "" for v in rows[-1]):
            rows.pop()
        return rows

    def parse_report_to_dataframe(self, data, columns=None, form_id=None):
        """将报表数据解析为 DataFrame"""
        if isinstance(data, dict):
            return {sheet_name: self.parse_report_to_dataframe(rows, columns=None, form_id=form_id) for sheet_name, rows in data.items()}

        if not isinstance(data, list) or len(data) == 0:
            if columns:
                return pd.DataFrame(columns=columns)
            return pd.DataFrame()

        if columns:
            df_columns = list(columns)
            if isinstance(data[0], (list, tuple)) and len(data[0]) > len(df_columns):
                df_columns.extend([f"__extra_field_{i}" for i in range(1, len(data[0]) - len(columns) + 1)])
            df = pd.DataFrame(data, columns=df_columns)
            output_columns = list(columns)
        else:
            df = pd.DataFrame(data)
            output_columns = None

        if form_id == "KDS_ReportData":
            return df

        # 过滤 AP/AR 汇总表中的“小计/合计”等行（小计可能出现在前两列）
        if form_id in ("AP_SumReport", "AR_SumReport") and len(df.columns) >= 2:
            first_col = df.columns[0]
            second_col = df.columns[1]
            for col in (first_col, second_col):
                df = df[~(df[col].astype(str).str.contains("小计", na=False, regex=False))]
                df = df[~(df[col].astype(str).str.contains("合计", na=False, regex=False))]

        for col in df.columns:
            if "日期" in str(col) or "Date" in str(col):
                try:
                    df[col] = pd.to_datetime(df[col], errors="coerce")
                    df[col] = df[col].dt.strftime("%Y-%m-%d")
                except Exception:
                    pass

        exclude_columns = []
        if form_id in ("AP_SumReport", "AR_SumReport") and len(df.columns) >= 3:
            exclude_columns = [df.columns[0], df.columns[1], df.columns[2]]
        if form_id == "HS_NoDimInOutStockDetailRpt" and len(df.columns) >= 1:
            if df.columns[0] not in exclude_columns:
                exclude_columns.append(df.columns[0])
        df = self._coerce_numeric_like_columns(df, exclude_columns=exclude_columns)

        if form_id in ("AP_SumReport", "AR_SumReport"):
            numeric_candidates = df.columns[3:]
        elif form_id == "HS_INOUTSTOCKSUMMARYRPT":
            numeric_candidates = [c for c in df.columns if c not in ["物料编码", "物料名称", "物料分组", "仓库"]]
        elif form_id == "HS_NoDimInOutStockDetailRpt":
            numeric_candidates = [c for c in df.columns if c not in ["期间", "单据日期", "单据编号", "业务类型", "单据类型", "物料编码", "物料名称"]]
        elif form_id in ("SAL_OutStockInvoiceRpt", "PUR_PurchaseOrderDetailRpt"):
            numeric_candidates = [
                c
                for c in df.columns
                if c
                not in [
                    "销售组织",
                    "单据编号",
                    "单据类型",
                    "日期",
                    "销售员",
                    "客户名称",
                    "物料名称",
                    "是否赠品",
                    "采购组织",
                    "订单编号",
                    "供应商名称",
                    "交货日期",
                    "结算币别",
                ]
            ]
        elif form_id == "GL_RPT_AccountBalance":
            numeric_candidates = [c for c in df.columns if c not in ["科目编码", "科目名称", "核算维度编码", "核算维度名称"]]
            if "科目编码" in df.columns:
                df["科目编码"] = df["科目编码"].map(self._format_account_code)
        else:
            numeric_candidates = []

        for col in numeric_candidates:
            if col in df.columns and not pd.api.types.is_numeric_dtype(df[col]):
                try:
                    df[col] = self._clean_numeric_series(df[col])
                except Exception:
                    pass

        if form_id == "PUR_PurchaseOrderDetailRpt":
            df = self._fill_purchase_order_detail_merged_headers(df)

        if output_columns:
            df = df[[col for col in output_columns if col in df.columns]]

        return df

    def _format_account_code(self, value):
        if pd.isna(value):
            return ""
        text = str(value).strip()
        if text.endswith(".0"):
            text = text[:-2]
        return text

    def _fill_purchase_order_detail_merged_headers(self, df):
        if df is None or df.empty or "订单编号" not in df.columns:
            return df

        header_columns = ["采购组织", "订单编号", "日期", "供应商名称"]
        header_columns = [col for col in header_columns if col in df.columns]
        if not header_columns:
            return df

        normalized_order_no = df["订单编号"].astype(str).str.strip()
        detail_value_columns = [
            col
            for col in ["物料名称", "订货数量", "价税合计", "收料数量", "入库数量", "应付数量", "开票数量", "已结算金额"]
            if col in df.columns
        ]

        if detail_value_columns:
            has_detail_values = df[detail_value_columns].astype(str).apply(lambda s: s.str.strip()).ne("").any(axis=1)
        else:
            has_detail_values = pd.Series(True, index=df.index)

        fill_mask = normalized_order_no.eq("") & has_detail_values
        filled_headers = df[header_columns].replace(r"^\s*$", pd.NA, regex=True).ffill()
        df.loc[fill_mask, header_columns] = filled_headers.loc[fill_mask, header_columns]
        return df

    def _clean_numeric_series(self, series):
        cleaned = self._normalize_numeric_text(series)
        return pd.to_numeric(cleaned, errors="coerce").fillna(0)

    def _normalize_numeric_text(self, series):
        if series is None:
            return series

        cleaned = series.astype(str)
        cleaned = cleaned.str.replace("\u00A0", "", regex=False)
        cleaned = cleaned.str.replace("\u3000", "", regex=False)
        cleaned = cleaned.str.replace(r"\s+", "", regex=True)
        cleaned = cleaned.str.strip()
        cleaned = cleaned.str.replace(",", "", regex=False)
        cleaned = cleaned.str.replace("，", "", regex=False)
        cleaned = cleaned.str.replace("￥", "", regex=False)
        cleaned = cleaned.str.replace("¥", "", regex=False)
        cleaned = cleaned.str.replace("$", "", regex=False)
        cleaned = cleaned.str.replace(r"^\((.*)\)$", r"-\1", regex=True)
        cleaned = cleaned.str.replace(r"^（(.*)）$", r"-\1", regex=True)
        cleaned = cleaned.str.replace(r"^(\d+(?:\.\d+)?)-$", r"-\1", regex=True)

        cleaned = cleaned.replace(
            {
                "": None,
                "None": None,
                "nan": None,
                "NaN": None,
                "-": None,
                "—": None,
                "–": None,
            }
        )
        return cleaned

    def _coerce_numeric_like_columns(self, df, exclude_columns=None):
        if df is None or df.empty:
            return df

        exclude_columns = set(exclude_columns or [])
        for col in df.columns:
            if col in exclude_columns:
                continue
            if pd.api.types.is_numeric_dtype(df[col]):
                continue
            if not pd.api.types.is_object_dtype(df[col]):
                continue

            series = df[col]
            cleaned = self._normalize_numeric_text(series)
            non_empty = cleaned.dropna()
            if non_empty.empty:
                continue

            numeric = pd.to_numeric(non_empty, errors="coerce")
            if numeric.notna().mean() >= 0.7:
                df[col] = self._clean_numeric_series(series)

        return df

    def build_fund_position_output(self, df):
        column_map = {
            "银行": "银行",
            "原币期初余额": "(原币)期初余额",
            "原币本日收入": "(原币)本期收入",
            "原币本日支出": "(原币)本期支出",
            "原币本日余额": "(原币)本期余额",
            "收入笔数": "收入笔数",
            "支出笔数": "支出笔数",
        }

        output_columns = list(column_map.values())
        if df is None or df.empty:
            return pd.DataFrame(columns=output_columns)

        df_selected = pd.DataFrame()
        for src_col, dst_col in column_map.items():
            if src_col in df.columns:
                df_selected[dst_col] = df[src_col]
            else:
                df_selected[dst_col] = 0

        numeric_cols = [col for col in df_selected.columns if col != "银行"]
        for col in numeric_cols:
            df_selected[col] = self._clean_numeric_series(df_selected[col])

        summary = {col: 0 for col in df_selected.columns}
        summary["银行"] = "合计"
        for col in numeric_cols:
            summary[col] = df_selected[col].sum()

        df_selected = pd.concat([df_selected, pd.DataFrame([summary])], ignore_index=True)

        for col in ["收入笔数", "支出笔数"]:
            if col in df_selected.columns:
                df_selected[col] = pd.to_numeric(df_selected[col], errors="coerce").fillna(0).astype(int)

        return df_selected[output_columns]

    def parse_data_to_dataframe(self, data, columns):
        """将数据解析为DataFrame，添加列名"""
        if not isinstance(data, list) or len(data) == 0:
            return pd.DataFrame(columns=columns)

        df_columns = list(columns)
        if isinstance(data[0], (list, tuple)) and len(data[0]) > len(df_columns):
            df_columns.extend([f"__extra_field_{i}" for i in range(1, len(data[0]) - len(columns) + 1)])
        df = pd.DataFrame(data, columns=df_columns)

        date_columns = ["日期", "业务日期", "采购日期", "申请日期", "要货日期", "交货日期", "到期日"]
        for col in date_columns:
            if col in df.columns:
                try:
                    df[col] = pd.to_datetime(df[col], errors="coerce")
                    df[col] = df[col].dt.strftime("%Y-%m-%d")
                except Exception as e:
                    print(f"  [WARN] 日期列 {col} 格式化失败: {e}")

        if "业务类型" in df.columns:
            business_type_map = {
                "NORMAL": "普通应付单",
                "DEFECT": "退货应付单",
                "CONSIGNMENT": "受托代销应付单",
                "COMMISSION": "委托代销应付单",
                "OTHERRECE": "其他应收单",
                "OTHERPAY": "其他应付单",
                "INITIALRECE": "期初应收单",
                "INITIALPAY": "期初应付单",
                "CG": "采购",
                "FY": "费用",
                "采购": "采购",
                "费用": "费用",
            }
            df["业务类型"] = df["业务类型"].map(lambda x: business_type_map.get(str(x), str(x)))

        if "往来单位类型" in df.columns:
            contact_type_map = {
                "BD_Customer": "客户",
                "BD_Supplier": "供应商",
                "HSWD01_SYS": "供应商",
                "ORG_Organizations": "组织机构",
                "BD_OtherOrg": "其他组织",
                "BD_Employee": "员工",
                "BD_Empinfo": "员工",
                "BD_BANK": "银行",
            }
            df["往来单位类型"] = df["往来单位类型"].map(lambda x: contact_type_map.get(str(x), str(x)))

        status_map = {
            "Z": "暂存",
            "A": "创建",
            "B": "审核中",
            "C": "已审核",
            "D": "重新审核",
        }
        close_status_map = {
            "A": "未关闭",
            "B": "已关闭",
        }
        giveaway_map = {
            "true": "是",
            "false": "否",
        }

        if "单据状态" in df.columns:
            df["单据状态"] = df["单据状态"].map(lambda x: status_map.get(str(x), str(x)))
        if "关闭状态" in df.columns:
            df["关闭状态"] = df["关闭状态"].map(lambda x: close_status_map.get(str(x), str(x)))
        for col in ["是否赠品", "申请借款", "实报实付"]:
            if col in df.columns:
                df[col] = df[col].map(lambda x: giveaway_map.get(str(x).lower(), str(x)))

        return df[[col for col in columns if col in df.columns]]

    def save_all_to_excel(self, dataframes_dict):
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"云星空经营数据_{self.period_name}_{timestamp}.xlsx"

        try:
            with pd.ExcelWriter(filename, engine="openpyxl") as writer:
                for sheet_name, df in dataframes_dict.items():
                    write_header = not self._is_kds_output_sheet(sheet_name)
                    df.to_excel(writer, sheet_name=sheet_name, index=False, header=write_header)

                    try:
                        ws = writer.sheets[sheet_name]
                        self._apply_excel_number_formats(sheet_name, df, ws)
                    except Exception as e:
                        print(f"  ! sheet {sheet_name} 数值格式设置失败: {e}")

            print(f"  Excel文件已生成: {filename}")
            return filename
        except Exception as e:
            print(f"  保存Excel失败: {e}")
            return None

    def _is_kds_output_sheet(self, sheet_name):
        return sheet_name in {"资产负债表", "利润表", "现金流量表"} or str(sheet_name).startswith("报表")

    def _apply_excel_number_formats(self, sheet_name, df, ws):
        try:
            from openpyxl.styles import numbers
        except Exception:
            return

        if df is None or df.empty:
            return

        if sheet_name == "资金头寸表":
            amount_cols = ["(原币)期初余额", "(原币)本期收入", "(原币)本期支出", "(原币)本期余额"]
            count_cols = ["收入笔数", "支出笔数"]
        else:
            amount_cols = []
            count_cols = []

        if sheet_name == "科目余额表" and "科目编码" in df.columns:
            col_idx = list(df.columns).index("科目编码") + 1
            for row in range(2, 2 + len(df)):
                cell = ws.cell(row=row, column=col_idx)
                cell.value = "" if cell.value is None else str(cell.value)
                cell.number_format = "@"

        if self._is_kds_output_sheet(sheet_name):
            for row in range(5, ws.max_row + 1):
                for col in range(2, ws.max_column + 1):
                    cell = ws.cell(row=row, column=col)
                    if cell.value is None or str(cell.value).strip() == "":
                        continue
                    try:
                        cell.value = float(str(cell.value).replace(",", "").strip())
                        cell.number_format = "#,##0.00"
                    except Exception:
                        pass

        for col_idx, col_name in enumerate(df.columns, start=1):
            series = df[col_name]
            if pd.api.types.is_integer_dtype(series):
                fmt = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
            elif pd.api.types.is_float_dtype(series):
                fmt = "#,##0.00"
            else:
                fmt = None

            if sheet_name == "资金头寸表":
                if col_name in amount_cols:
                    fmt = "#,##0.00"
                if col_name in count_cols:
                    fmt = numbers.FORMAT_NUMBER_COMMA_SEPARATED1

            if not fmt:
                continue

            for row in range(2, 2 + len(df)):
                ws.cell(row=row, column=col_idx).number_format = fmt

    def _drop_rows_with_empty_contactunit_name(self, df, contactunit_name_col="往来单位名称"):
        if df is None or df.empty:
            return df
        if contactunit_name_col not in df.columns:
            return df
        series = df[contactunit_name_col]
        return df[series.notna() & (series.astype(str).str.strip() != "")]

    def send_summary_with_file(self, excel_file, bill_records):
        if self.no_wechat:
            print("  已跳过企业微信推送（--no-wechat）")
            return True
        if not self.wechat_webhook:
            print("  已跳过企业微信推送（WECHAT_CONFIG.webhook 为空）")
            return True

        try:
            file_size = os.path.getsize(excel_file)
            file_size_mb = file_size / (1024 * 1024)

            export_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

            summary_lines = [
                "云星空经营数据导出完成",
                "",
                f"数据周期: {self.period_name}",
                f"导出时间: {export_time}",
                "",
                "数据明细:",
            ]

            for idx, record in enumerate(bill_records, 1):
                summary_lines.append(f"  {idx}. {record['name']}: {record['count']} 条")

            summary_lines.append("")
            summary_lines.append("文件信息:")
            summary_lines.append(f"  文件名: {os.path.basename(excel_file)}")
            summary_lines.append(f"  大小: {file_size_mb:.2f} MB")

            summary_msg = "\n".join(summary_lines)

            payload = {"msgtype": "text", "text": {"content": summary_msg}}

            response = requests.post(self.wechat_webhook, json=payload, headers={"Content-Type": "application/json"})
            result = response.json()
            if result.get("errcode") == 0:
                print("  汇总消息发送成功")
                return True
            print(f"  汇总消息发送失败: {result}")
            return False
        except Exception as e:
            print(f"  发送汇总消息异常: {e}")
            return False

    def export_all_bills(self):
        print("=" * 60)
        print(f"销售单据数据导出 - {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        print("=" * 60)

        try:
            if not self.login_kingdee():
                return False

            self._resolve_org_scope_after_login()

            all_dataframes = {}
            bill_records = []

            for config in self.bill_configs:
                form_id = config["form_id"]
                bill_name = config["bill_name"]
                if self.only and (form_id.lower() not in self.only) and (bill_name.lower() not in self.only):
                    continue
                field_keys = self._append_extra_field_keys(config, config["field_keys"])
                filter_string = config["filter_string"]
                columns = config["columns"]

                print(f"\n处理单据: {bill_name}")
                print("-" * 60)

                data = self.get_bill_data_with_filter(form_id, field_keys, filter_string)
                df = self.parse_data_to_dataframe(data, columns)

                record_count = len(df)
                bill_records.append({"name": bill_name, "count": record_count})
                all_dataframes[bill_name] = df

                if record_count == 0:
                    print(f"  [WARN] {bill_name} 无数据（将创建空表）")
                else:
                    print(f"  {bill_name} 获取到 {record_count} 条记录")

            for config in self.report_configs:
                form_id = config["form_id"]
                report_name = config["report_name"]
                if self.only and (form_id.lower() not in self.only) and (report_name.lower() not in self.only):
                    continue
                field_keys = self._append_extra_field_keys(config, config.get("field_keys", ""))
                model = copy.deepcopy(config["model"])
                columns = config["columns"]

                if form_id in ("AP_SumReport", "AR_SumReport"):
                    settle_org_lst = self.resolve_settle_org_ids_by_numbers(self.target_settle_org_numbers, self.default_settle_org_id_map)
                    if form_id == "AP_SumReport":
                        model = self.build_ap_sum_report_model(settle_org_lst)
                    else:
                        model = self.build_ar_sum_report_model(settle_org_lst)
                elif config.get("org_id_model_field"):
                    settle_org_lst = self.resolve_settle_org_ids_by_numbers(self.target_settle_org_numbers, self.default_settle_org_id_map)
                    model[config["org_id_model_field"]] = settle_org_lst
                elif config.get("api_type") == "kds_report":
                    model["OrgNumber"] = self.inventory_org_number or (self.target_settle_org_numbers[0] if self.target_settle_org_numbers else "")

                print(f"\n处理报表: {report_name}")
                print("-" * 60)

                if config.get("api_type") == "kds_report":
                    data = self.get_kds_report_data(model)
                else:
                    data = self.get_report_data(form_id, field_keys, model, scheme_id=config.get("scheme_id", ""))
                df = self.parse_report_to_dataframe(data, columns, form_id=form_id)

                if isinstance(df, dict):
                    total_count = 0
                    for sheet_name, sheet_df in df.items():
                        record_count = len(sheet_df)
                        total_count += record_count
                        bill_records.append({"name": sheet_name, "count": record_count})
                        all_dataframes[sheet_name] = sheet_df
                        if record_count == 0:
                            print(f"  [WARN] {sheet_name} 无数据（将创建空表）")
                        else:
                            print(f"  {sheet_name} 获取到 {record_count} 条记录")
                    continue

                if form_id in ("AR_SumReport", "AP_SumReport"):
                    df = self._drop_rows_with_empty_contactunit_name(df, contactunit_name_col="往来单位名称")

                if report_name == "资金头寸表":
                    df = self.build_fund_position_output(df)

                record_count = len(df)
                bill_records.append({"name": report_name, "count": record_count})
                all_dataframes[report_name] = df

                if record_count == 0:
                    print(f"  [WARN] {report_name} 无数据（将创建空表）")
                else:
                    print(f"  {report_name} 获取到 {record_count} 条记录")

            print("\n正在生成Excel文件...")
            excel_file = self.save_all_to_excel(all_dataframes)

            if not excel_file:
                print("  Excel生成失败")
                return False

            self.send_summary_with_file(excel_file, bill_records)

            print("\n生成的文件:")
            if os.path.exists(excel_file):
                print(f"  - {os.path.abspath(excel_file)}")

            print("=" * 60)
            print("导出任务完成")
            print("=" * 60)

            return True

        except Exception as e:
            print(f"导出任务异常: {e}")
            import traceback

            traceback.print_exc()
            return False


def main():
    start_date = None
    end_date = None
    no_wechat = False
    org_numbers = None
    only = None
    extra_fields = None
    list_orgs = False
    show_config = False

    if "--start" in sys.argv:
        try:
            start_date = sys.argv[sys.argv.index("--start") + 1]
        except Exception:
            start_date = None
    if "--end" in sys.argv:
        try:
            end_date = sys.argv[sys.argv.index("--end") + 1]
        except Exception:
            end_date = None
    if "--no-wechat" in sys.argv:
        no_wechat = True
    if "--org" in sys.argv:
        try:
            org_numbers = sys.argv[sys.argv.index("--org") + 1]
        except Exception:
            org_numbers = None
    if "--only" in sys.argv:
        try:
            only = sys.argv[sys.argv.index("--only") + 1]
        except Exception:
            only = None
    if "--fields" in sys.argv:
        try:
            extra_fields = sys.argv[sys.argv.index("--fields") + 1]
        except Exception:
            extra_fields = None
    if "--list-orgs" in sys.argv:
        list_orgs = True
    if "--show-config" in sys.argv:
        show_config = True

    exporter = SalesDataExporter(
        start_date=start_date,
        end_date=end_date,
        no_wechat=no_wechat,
        org_numbers=org_numbers,
        only=only,
        extra_fields=extra_fields,
    )

    if show_config:
        print("可用导出项（--only 可填 form_id 或名称，支持逗号分隔）：")
        print("-" * 60)
        for c in exporter.bill_configs:
            print(f"[BILL] {c['form_id']}  |  {c['bill_name']}")
        for c in exporter.report_configs:
            print(f"[RPT ] {c['form_id']}  |  {c['report_name']}")
        return

    if list_orgs:
        orgs = exporter.get_all_organizations()
        try:
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"组织列表_{ts}.xlsx"
            pd.DataFrame(orgs).to_excel(filename, index=False)
            print(f"组织列表已导出：{os.path.abspath(filename)}")
        except Exception as e:
            print(f"组织列表导出失败（将改为打印到控制台）：{e}")
            for o in orgs:
                print(f"{o.get('number')}\t{o.get('name')}")
        return

    success = exporter.export_all_bills()
    if not success:
        sys.exit(1)


if __name__ == "__main__":
    main()

